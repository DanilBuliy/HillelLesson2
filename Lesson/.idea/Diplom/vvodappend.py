import openpyxl
import datetime
import openpyxl
import datetime

def questdata():
    while True:
        pol = input("Пол(м/ж): ")
        if not pol.isalpha():
            print("Непонятный ввод")
            continue
        else:
            break

    while True:
        firstname = input("Введите имя: ")
        if not firstname.isalpha():
            print("Это не похоже на имя")
        else:
            break

    while True:
        secondname = input("Введите фамилию: ")
        if secondname == "":
            secondname = "-"
            break
        elif not secondname.isalpha():
            print("Это не похоже на фамилию")
        else:
            break

    while True:
        strotec = input("Введите Отчество: ")
        if strotec == "":
            strotec = "-"
            break
        elif not strotec.isalpha():
            print("Это не похоже на Отчество")
        else:
            break



    while True:
        birthday = input("Введите дату рождения (число/месяц/год): ")
        if (birthday.isalpha() or not birthday.replace(".", "").replace("/", "").replace(" ", "").replace("-","").isdigit()
                or "." in birthday[-1] or "/" in birthday[-1] or "-" in birthday[-1]):
            print("непонятный ввод")
            continue

        if "/" in birthday:
            sbirthday = birthday.split("/")
        elif "." in birthday:
            sbirthday = birthday.split(".")
        elif "-" in birthday:
            sbirthday = birthday.split("-")
        else:
            sbirthday = birthday.split()

        try:
            sbirthday = list(map(int, sbirthday))
        except Exception:
            print("Ошибка ввода")

        if len(sbirthday) < 3:
            print("Не всю дату ввели")
            continue

        chislo = sbirthday[0]
        mes = sbirthday[1]
        year = sbirthday[2]

        nowyear = datetime.date.today().year
        nowmes = datetime.date.today().month
        nowday = datetime.date.today().day

        if not 1 <= mes <= 12:
            print("Месяц введено не верно")
            continue
        if not 1 <= chislo <= 31:
            print("Число введено не верно")
            continue
        if year > nowyear:
            print("Год введено не верно")
            continue
        if nowmes == mes and chislo >= nowday:
            age = nowyear - year
        else:
            age = nowyear - year - 1
        break

    while True:
        deathday = input("Введите дату смерти (если присутствует) \nЕсли отсутствует - ENTER: ")
        if deathday == "":
            deathday = "-"
            break

        if (deathday.isalpha() or not deathday.replace(".", "").replace("/", "").replace(" ", "").replace("-","").isdigit()
                or "." in deathday[-1] or "/" in deathday[-1] or "-" in deathday[-1]):
            print("Непонятный ввод")
            continue

        if "/" in deathday:
            sdeathday = deathday.split("/")
        elif "." in deathday:
            sdeathday = deathday.split(".")
        elif "-" in deathday:
            sdeathday = deathday.split("-")
        else:
            sdeathday = deathday.split()

        try:
            sdeathday = list(map(int, sdeathday))
        except Exception:
            print("Ошибка ввода")
        else:
            if len(sdeathday) < 3:
                print("Ошибка ввода")
                continue

            dchislo = sdeathday[0]
            dmes = sdeathday[1]
            dyear = sdeathday[2]

            if not 1 <= dmes <= 12:
                print("Месяц введено не верно")
                continue

            if not 1 <= dchislo <= 31:
                print("Число введено не верно")
                continue

            if dyear < year:
                print("Человек не мог умереть раньше, чем родился")
                continue

            if dyear > nowyear:
                print("Год введено не верно")
                continue

            if dmes < mes or dmes == mes and dchislo < chislo:
                age = dyear - year - 1
            else:
                age = dyear - year
            break

    if age % 10 == 1:
        age_str = f"{age} год"
    elif age % 10 in [2, 3, 4]:
        age_str = f"{age} года"
    else:
        age_str = f"{age} лет"

    values = [pol, firstname, secondname, strotec, f"{birthday}, ({age_str})", deathday]

    add = ["Пол", "Имя", "Фамилия", "Отчество", "Дата рождения", "Дата смерти"]
    print(*add)
    print(*values)

    return values
def vvod():
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    add = ["Пол", "Имя", "Фамилия","Отчество", "Дата рождения", "Дата смерти"]
    for ind, i in enumerate(add, start=1):
        sheet1.cell(row=1, column=ind, value=i)

    values=questdata()
    sheet1.append(values)

    workbook.save("data.xlsx")

def vvod2():
    file_path = "data.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet1 = workbook.active
    except FileNotFoundError:
        print("Такого файла нет, создаем новый")
        workbook = openpyxl.Workbook()
        sheet1 = workbook.active
        add = ["Пол", "Имя", "Фамилия", "Отчество", "Дата рождения", "Дата смерти"]
        for ind, i in enumerate(add, start=1):
            sheet1.cell(row=1, column=ind, value=i)

    values=questdata()
    sheet1.append(values)

    workbook.save(file_path)